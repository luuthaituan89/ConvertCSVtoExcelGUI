import pandas as pd
import csv
import os
from tqdm import tqdm
from openpyxl import load_workbook

def get_valid_path(prompt, check_dir=False):
    """Function to get a valid file or directory path from user."""
    while True:
        path = input(prompt)
        if os.path.exists(path) and (not check_dir or os.path.isdir(path)):
            return path
        print(f"Error: The {'directory' if check_dir else 'file'} does not exist. Please try again.")

# Yêu cầu người dùng nhập đường dẫn file
file_path = get_valid_path("Please enter the path to the CSV file: ")

if not file_path.lower().endswith('.csv'):
    print("Error: The file is not in CSV format.")
else:
    try:
        # Tự động xác định delimiter bằng Sniffer
        with open(file_path, mode='r', encoding='utf-8') as file:
            dialect = csv.Sniffer().sniff(file.read(1024))
            file.seek(0)

            # Yêu cầu người dùng nhập đường dẫn thư mục lưu
            output_dir = get_valid_path("Please enter the directory to save the Excel file: ", check_dir=True)

            # Yêu cầu người dùng nhập tên file
            output_file_name = input("Please enter the name for the Excel file (without extension): ")
            output_file = os.path.join(output_dir, f"{output_file_name}.xlsx")

            # Xác định kích thước chunk và số dòng tổng
            chunk_size = 10000
            total_lines = sum(1 for _ in open(file_path, 'r', encoding='utf-8')) - 1
            num_chunks = (total_lines // chunk_size) + 1

            print("Converting CSV to Excel...")

            # Ghi từng chunk vào file Excel
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                with tqdm(total=100, desc="Processing", unit="%", unit_scale=True) as pbar:
                    for i, chunk in enumerate(pd.read_csv(file, delimiter=dialect.delimiter, chunksize=chunk_size)):
                        startrow = writer.sheets['Sheet1'].max_row if 'Sheet1' in writer.sheets else 0
                        chunk.to_excel(writer, index=False, header=writer.sheets.get('Sheet1') is None, startrow=startrow)
                        pbar.update((i + 1) * 100 / num_chunks - pbar.n)

            # Điều chỉnh độ rộng cột
            wb = load_workbook(output_file)
            ws = wb.active

            print("Adjusting column widths...")
            for col in tqdm(ws.columns, desc="Processing columns", unit="col"):
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(output_file)
            print("Conversion complete. File saved as", output_file)
    except Exception as e:
        print("An error occurred:", e)
