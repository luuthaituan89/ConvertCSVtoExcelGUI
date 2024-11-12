import pandas as pd
import csv
import os
from tkinter import Label, Entry, Button, filedialog, messagebox, StringVar, Checkbutton, BooleanVar, PhotoImage
from tkinter.ttk import Progressbar, Frame, Style
from ttkthemes import ThemedTk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import threading
import logging

# Configure logging
logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def select_csv_file():
    try:
        file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        csv_file_path.set(file_path)
        logging.info(f'Selected CSV file: {file_path}')
    except Exception as e:
        logging.error(f'Error selecting CSV file: {e}')

def select_output_dir():
    try:
        dir_path = filedialog.askdirectory()
        output_dir_path.set(dir_path)
        logging.info(f'Selected output directory: {dir_path}')
    except Exception as e:
        logging.error(f'Error selecting output directory: {e}')

def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

def validate_inputs(file_path, output_dir, output_file_name):
    if not file_path or not output_dir or not output_file_name:
        messagebox.showerror("Error", "Please complete all fields.")
        logging.warning('Validation failed: Missing fields')
        return False
    if not file_path.lower().endswith('.csv'):
        messagebox.showerror("Error", "The selected file is not in CSV format.")
        logging.warning('Validation failed: File is not CSV')
        return False
    if not os.path.isfile(file_path):
        messagebox.showerror("Error", "The CSV file does not exist.")
        logging.warning('Validation failed: CSV file does not exist')
        return False
    if not os.path.isdir(output_dir):
        messagebox.showerror("Error", "The output directory does not exist.")
        logging.warning('Validation failed: Output directory does not exist')
        return False
    if any(char in output_file_name for char in r'\/:*?"<>|'):
        messagebox.showerror("Error", "The file name contains invalid characters.")
        logging.warning('Validation failed: Invalid characters in file name')
        return False
    return True

def convert_csv_to_excel():
    logging.info('Starting CSV to Excel conversion')
    status_label_var.set("Converting...")
    convert_button.config(state="disabled")
    root.update_idletasks()

    file_path = csv_file_path.get()
    output_dir = output_dir_path.get()
    output_file_name = name_entry.get().strip()

    if not validate_inputs(file_path, output_dir, output_file_name):
        convert_button.config(state="normal")
        return

    output_file = os.path.join(output_dir, f"{output_file_name}.xlsx")
    try:
        with open(file_path, mode='r', encoding='utf-8') as file:
            dialect = csv.Sniffer().sniff(file.read(1024))
            file.seek(0)

            # Get total lines and chunks
            total_lines = sum(1 for _ in open(file_path, 'r', encoding='utf-8')) - 1
            chunk_size = 10000
            num_chunks = (total_lines // chunk_size) + 1

            # Progress calculation setup
            total_steps = num_chunks + 1  # +1 for column adjustment
            progress_increment = 100 / total_steps

            progress_bar['value'] = 0
            root.update_idletasks()

            sheet_index = 1
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                if not split_checkbox_var.get():
                    chunk = pd.read_csv(file, delimiter=dialect.delimiter)
                    chunk.to_excel(writer, index=False, sheet_name='Sheet1')
                    progress_bar['value'] += progress_increment
                    root.update_idletasks()
                else:
                    for i, chunk in enumerate(pd.read_csv(file, delimiter=dialect.delimiter, chunksize=chunk_size)):
                        startrow = 0 if writer.sheets.get(f'Sheet{sheet_index}') is None else writer.sheets[
                            f'Sheet{sheet_index}'].max_row

                        if startrow + len(chunk) > 1048576:
                            sheet_index += 1
                            startrow = 0

                        sheet_name = f'Sheet{sheet_index}'
                        chunk.to_excel(writer, index=False, header=writer.sheets.get(sheet_name) is None,
                                       startrow=startrow, sheet_name=sheet_name)
                        progress_bar['value'] += progress_increment
                        root.update_idletasks()

            # Adjust column widths
            status_label_var.set("Adjusting columns...")
            root.update_idletasks()
            wb = load_workbook(output_file)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                adjust_column_widths(ws)
                progress_bar['value'] += progress_increment / len(wb.sheetnames)
                root.update_idletasks()
            wb.save(output_file)

            status_label_var.set("Completed")
            progress_bar['value'] = 100
            logging.info(f'Conversion completed successfully, output saved to {output_file}')
            messagebox.showinfo("Success", f"Conversion complete. File saved as {output_file}")

    except Exception as e:
        logging.error(f'Error during conversion: {e}')
        messagebox.showerror("Error", f"An error occurred: {e}")
        status_label_var.set("Error")

    finally:
        convert_button.config(state="normal")

def start_conversion_thread():
    threading.Thread(target=convert_csv_to_excel, daemon=True).start()

def on_closing():
    root.quit()
    root.destroy()

# Setup GUI
root = ThemedTk(theme="winnative")
root.title("CSV to Excel Converter")

# Window dimensions
window_width = 650
window_height = 300
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
position_top = int(screen_height / 2 - window_height / 2)
position_right = int(screen_width / 2 - window_width / 2)
root.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")
root.resizable(False, False)

# Icon
try:
    root.iconphoto(False, PhotoImage(file="icon.png"))
except:
    pass

# Style
style = Style()
style.configure("TButton", font=("Arial", 12))
style.configure("TLabel", font=("Arial", 10))
style.configure("TEntry", font=("Arial", 10))

# Variables
csv_file_path = StringVar()
output_dir_path = StringVar()
status_label_var = StringVar()
split_checkbox_var = BooleanVar()

# Frames
main_frame = Frame(root, padding=(20, 20))
main_frame.pack(fill="both", expand=True)
main_frame.grid_columnconfigure(1, weight=1)

# Widgets
Label(main_frame, text="Select CSV file:").grid(row=0, column=0, sticky="w", pady=5, padx=5)
csv_entry = Entry(main_frame, textvariable=csv_file_path, width=50)
csv_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
Button(main_frame, text="Browse", command=select_csv_file).grid(row=0, column=2, padx=5, pady=5)

Label(main_frame, text="Select output directory:").grid(row=1, column=0, sticky="w", pady=5, padx=5)
dir_entry = Entry(main_frame, textvariable=output_dir_path, width=50)
dir_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
Button(main_frame, text="Browse", command=select_output_dir).grid(row=1, column=2, padx=5, pady=5)

Label(main_frame, text="Enter Excel file name (without extension):").grid(row=2, column=0, sticky="w", pady=5, padx=5)
name_entry = Entry(main_frame, width=50)
name_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

split_checkbox = Checkbutton(main_frame, text="Split CSV into multiple sheets (if needed)", variable=split_checkbox_var)
split_checkbox.grid(row=3, column=1, pady=10, sticky="w")

convert_button = Button(main_frame, text="Convert to Excel", command=start_conversion_thread)
convert_button.grid(row=4, column=1, pady=20)

status_label = Label(main_frame, textvariable=status_label_var, font=("Arial", 10))
status_label.grid(row=5, column=1, pady=5)

progress_bar = Progressbar(main_frame, orient="horizontal", length=300, mode="determinate")
progress_bar.grid(row=6, column=1, pady=10, sticky="ew")

root.protocol("WM_DELETE_WINDOW", on_closing)
root.mainloop()
